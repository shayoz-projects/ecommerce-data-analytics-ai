

import sqlite3

import pandas as pd
from openpyxl import load_workbook


def load_orders_xlsx(file_path):

    workbook = load_workbook(file_path, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    columns = rows[0]
    data = rows[1:]

    return pd.DataFrame(data, columns=columns)


orders_2023_2024 = load_orders_xlsx(
    "Orders_products_-2023-2024.xlsx"

)

orders_2024_2026 = load_orders_xlsx(
    "Orders__products_2024-2026.xlsx"
)

orders = pd.concat(
    [orders_2023_2024, orders_2024_2026],
    ignore_index=True
)


# convert data 
orders["Date created"] = pd.to_datetime(
    orders["Date created"],
    errors="coerce",
)

orders["Price"] = pd.to_numeric(
    orders["Price"],
    errors="coerce",
)

orders["Qty"] = pd.to_numeric(
    orders["Qty"],
    errors="coerce",
)

orders["line_revenue"] = orders["Price"] * orders["Qty"]

orders["email_norm"] = (
    orders["Contact email"]
    .str.strip()
    .str.lower()
)

orders["is_gift_voucher"] = (
    orders["Item"]
    .str.contains("שובר", na=False)
)

# canceled orders 
orders["is_canceled"] = (
    orders["Fulfillment status"] == "Canceled"
)



# normalize product names
product_mapping = pd.read_csv(
    "product_name_mapping_FINAL.csv"
)

orders = orders.merge(
    product_mapping,
    left_on="Item",
    right_on="raw_order_item",
    how="left",
)

orders["canonical_name"] = (
    orders["canonical_name"]
    .fillna(orders["Item"])
)

orders = orders.drop(columns=["raw_order_item"])


# Keep the fields that been used 
order_columns = [
    "Order number",
    "Date created",
    "Total order quantity",
    "Contact email",
    "email_norm",
    "Item",
    "canonical_name",
    "Variant",
    "Qty",
    "Price",
    "line_revenue",
    "Quantity refunded",
    "is_gift_voucher",
    "is_canceled",
    "Delivery city",
    "Delivery country",
    "Fulfillment status",
    "source_file",
]

orders_clean = (
    orders[order_columns]
    .sort_values("Order number")
    .reset_index(drop=True)
)


# Clean the product catalog
catalog = pd.read_csv("catalog_products.csv")

catalog_columns_to_fill = [
    "name",
    "collection",
    "price",
    "visible",
    "cost",
    "sku",
]

catalog[catalog_columns_to_fill] = (
    catalog.groupby("handleId")[catalog_columns_to_fill]
    .ffill()
)

catalog_clean = (
    catalog[
        [
            "handleId",
            "name",
            "collection",
            "price",
            "inventory",
            "visible",
        ]
    ]
    .drop_duplicates(subset="handleId")
    .reset_index(drop=True)
)


# Clean the contacts table
contacts = pd.read_csv("contacts.csv")

contacts["email_norm"] = (
    contacts["Email 1"]
    .str.strip()
    .str.lower()
)

contacts_clean = contacts.drop_duplicates(
    subset="email_norm",
    keep="first",
)

contacts_clean = contacts_clean[
    [
        "First Name",
        "Last Name",
        "email_norm",
        "Phone 1",
        "Source",
    ]
].reset_index(drop=True)


# Load the cleaned tables into SQLite
database_name = "dataP.db"

with sqlite3.connect(database_name) as connection:
    orders_clean.to_sql(
        "orders",
        connection,
        if_exists="replace",
        index=False,
    )

    catalog_clean.to_sql(
        "products",
        connection,
        if_exists="replace",
        index=False,
    )

    contacts_clean.to_sql(
        "contacts",
        connection,
        if_exists="replace",
        index=False,
    )


print(f"Data loaded into {database_name}")
print(f"Orders: {len(orders_clean)} rows")
print(f"Products: {len(catalog_clean)} rows")
print(f"Contacts: {len(contacts_clean)} rows")
